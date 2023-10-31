#m 가로, n 세로 곱셈표 만들기
'''
m*n표 1 2 3 4 5
  1 
  2
  3
  4
  5
  6
'''
import sys, openpyxl,os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
os.chdir('C:\\Users\\USER\\Desktop\\AlgorithmQuiz')

wb = openpyxl.Workbook()
sheet = wb.active

# sys로 값 입력받기
m,n = map(int,sys.stdin.readline().split()) 

# A1번칸에 'mxn표' 문자열 넣기
sheet['A1'] = f"{m}x{n}표"

#가로 M값 넣기
#B1 C1 D1...위치에
m_num = 1
for i in range(98, 98+m):
    # sheet[f'{chr(i)}1'].font = title_style
    # sheet[f'{chr(i)}1'].fill = title_color
    sheet[f'{chr(i)}1'] = m_num
    
    m_num += 1

#세로 N값 넣기
#A2 A3 A4...위치에
for j in range(2,n+2):
    # sheet[f'A{j}'].font = title_style
    # sheet[f'A{j}'].fill = title_color
    sheet[f'A{j}'] = j-1

#곱셈값 넣기
#row 세로, column 가로
for i in range(2,m+2): #가로
  for j in range(2,n+2): #세로
    #가로값을 A로 고정하고 세로를 1,2,3..순
    row_val = int(sheet.cell(row=j, column=1).value)
    #세로값을 1로 고정하고 가로를 A,B,C..순
    col_val = int(sheet.cell(row=1, column=i).value)
    #위에서 받은 값을 서로 곱해서 값을 넣기
    sheet.cell(row=j, column=i, value= row_val * col_val)

#가운데정렬, 텍스트 스타일, 셀 스타일 변경
title_style = Font(size=12, bold=True)
title_color = PatternFill(start_color='F8CBAD',fill_type='solid')
all_border = Border(top = Side(border_style='thin', color='000000'),
                    bottom = Side(border_style='thin', color='000000'),
                    left = Side(border_style='thin', color='000000'),
                    right = Side(border_style='thin', color='000000'),)

for i in sheet.rows:
   for cell in i:
      cell.alignment = Alignment(horizontal='center', vertical='center')
      cell.border = all_border
      
      if cell.column == 1 or cell.row == 1:
         cell.font = title_style
         cell.fill = title_color
wb.save('multiplication_table.xlsx')